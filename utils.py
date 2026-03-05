import os
import re
import pandas as pd
from openpyxl import load_workbook
from typing import List, Dict

def setup_directories(input_dir: str):
    if not os.path.exists(input_dir):
        os.makedirs(input_dir)
        print(f"Created input directory: {input_dir}")

def save_to_excel(data_list: List[Dict], filename: str):
    """
    Appends list of invoice data dictionaries to Excel.
    Creates file if it doesn't exist.
    """
    if not data_list:
        return

    df_new = pd.DataFrame(data_list)
    # Ensure column order
    cols = [
        "archivo", "estado", "proveedor", "nit", "factura_numero", 
        "fecha", "fecha_vencimiento", "descripcion", 
        "moneda", "base", "impuestos", "total", 
        "direccion", "telefono", "ciudad", "cufe", 
        "nit_dv_calculado", "nit_dv_extraido", "nota"
    ]
    # Add missing cols if any
    for c in cols:
        if c not in df_new.columns:
            df_new[c] = None
    df_new = df_new[cols]

    if not os.path.exists(filename):
        # Create new
        df_new.to_excel(filename, index=False)
        print(f"Created new Excel file: {filename}")
    else:
        # Append
        # Load existing to check basic structure or just append using openpyxl for speed?
        # Pandas append is deprecated, use concatenation or openpyxl directly.
        # Simplest consistent way: Read, Concat, Write (fine for small-medium files).
        # For true append without reading, we need openpyxl.
        
        try:
            with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                # We need to find the last row.
                # If sheet exists
                try:
                    writer.book[writer.book.sheetnames[0]]
                    start_row = writer.book[writer.book.sheetnames[0]].max_row
                    df_new.to_excel(writer, index=False, header=False, startrow=start_row)
                except KeyError:
                    # Sheet might not exist if file was corrupted or empty? 
                    # If this fails, fallback to simple write
                    df_new.to_excel(filename, index=False)
            print(f"Appended {len(data_list)} rows to {filename}")
            
        except Exception as e:
            print(f"Error appending to Excel (fallback to read components): {e}")
            # Fallback: Read all, append, write all
            df_old = pd.read_excel(filename)
            df_combined = pd.concat([df_old, df_new], ignore_index=True)
            df_combined.to_excel(filename, index=False)
            print(f"Recovered and saved {len(df_combined)} rows.")

def clean_colombian_number(value) -> float:
    """
    Parses a number string assuming Colombian format:
    1. Removes currency symbols ($, COP).
    2. Handles dots (.) as thousands separators.
    3. Handles commas (,) as decimal separators.
    4. Handles space as thousands separators in specific cases.
    Example: "33.613,45" -> 33613.45
    Example: "40.000" -> 40000.0
    """
    if value is None:
        return 0.0
    
    if isinstance(value, (int, float)):
        return float(value)
        
    text = str(value).strip()
    
    # Remove currency symbols and non-numeric chars (except . , -)
    text = re.sub(r'[^\d.,-]', '', text)
    
    if not text:
        return 0.0

    # Colombian Format Heuristic
    # Case 1: "33.613,45" -> standard format
    if '.' in text and ',' in text:
        text = text.replace('.', '').replace(',', '.')
    
    # Case 2: "40.000" -> ambiguous, but in COL usually means 40k
    elif text.count('.') > 0 and ',' not in text:
        text = text.replace('.', '')
        
    # Case 3: "40,00" -> Decimal
    elif ',' in text and '.' not in text:
        text = text.replace(',', '.')
        
    try:
        return float(text)
    except ValueError:
        return 0.0

def format_date_colombian(value) -> str:
    """
    Standardizes dates to DD/MM/YYYY string format.
    Accepts: YYYY-MM-DD, DD-MM-YYYY, etc.
    """
    if not value:
        return ""
        
    text = str(value).strip()
    
    # Basic ISO parsing (YYYY-MM-DD) which is common in XML
    iso_match = re.match(r'(\d{4})[-/](\d{2})[-/](\d{2})', text)
    if iso_match:
        y, m, d = iso_match.groups()
        return f"{d}/{m}/{y}"
        
    # Try parsing DMY
    # If the date is already DD/MM/YYYY (e.g. 31/01/2026), keep it.
    dmy_match = re.match(r'(\d{2})[-/](\d{2})[-/](\d{4})', text)
    if dmy_match:
        d, m, y = dmy_match.groups()
        return f"{d}/{m}/{y}"
    
    return text

def clean_nit(value) -> str:
    """
    Standardizes NIT or ID format.
    Removes dots, spaces.
    Keeps hyphens if present (often used for verification digit).
    Example: "900.123.456-1" -> "900123456-1"
    Example: "1.032.443.194" -> "1032443194"
    """
    if not value:
        return ""
        
    # Ensure string
    text = str(value).strip()
    
    # Remove dots and spaces
    text = text.replace(".", "").replace(" ", "")
    
    return text

def calculate_nit_verification_digit(nit_number: str) -> str:
    """
    Calculates the verification digit for a given NIT (without DV).
    Algorithm: Modulo 11 with specific weights.
    Weights: 41, 37, 29, 23, 19, 17, 13, 7, 3
    """
    if not nit_number or not nit_number.isdigit():
        return ""
        
    # Weights for up to 15 digits (from right to left)
    # But usually NITs are shorter. 
    # DIAN rule: 41 37 29 23 19 17 13 7 3
    # applied from right to left? No, usually applied to the number as a sequence.
    # The prompt says: "De izquierda a derecha, todas las cifras se multiplican por una secuencia 41 37 29 23 19 17 13 7 3"
    # Wait, usually for shorter numbers we align to the right? 
    # DIAN weights for up to 15 digits (from right to left: 3, 7, 13...)
    # But usually applied left-to-right on the number string.
    # We need to align the weights to the RIGHT of the number.
    # Primes: 3, 7, 13, 17, 19, 23, 29, 37, 41, 43, 47, 53, 59, 67, 71
    
    primes = [71, 67, 59, 53, 47, 43, 41, 37, 29, 23, 19, 17, 13, 7, 3]
    
    # Cleaning
    nit_str = nit_number.strip()
    
    # We only support up to 15 digits with this array
    if len(nit_str) > 15:
        return "?"
        
    # We need to take the LAST N weights where N = len(nit_str)
    # Example: if len is 9, we take last 9: 41, 37...
    # Example: if len is 10, we take last 10: 43, 41...
    
    current_weights = primes[-len(nit_str):]
    
    total = 0
    for i, digit in enumerate(nit_str):
        if not digit.isdigit():
             return "?"
        total += int(digit) * current_weights[i]
        
    v = total % 11
    if v > 1:
        v = 11 - v
    else:
        v = v 
        
    return str(v)

def correct_suspicious_low_total(value, threshold: float = 500.0):
    """
    Heuristic: values below threshold (e.g. 500 COP) are often decimal errors (missing *1000).
    Returns corrected value if 0 < value < threshold, else returns value unchanged.
    """
    if value is None or not isinstance(value, (int, float)):
        return value
    try:
        v = float(value)
        if 0 < v < threshold:
            return v * 1000
    except (TypeError, ValueError):
        pass
    return value


def deduplicate_invoice_rows(df: pd.DataFrame, key_column: str = "archivo") -> pd.DataFrame:
    """
    Removes duplicate rows by key_column (e.g. archivo), keeping the row with the most non-null values.
    """
    if df.empty or key_column not in df.columns:
        return df
    df = df.copy()
    df["_completeness"] = df.notna().sum(axis=1)
    df = df.sort_values(by=[key_column, "_completeness"], ascending=[True, False])
    df = df.drop_duplicates(subset=[key_column], keep="first")
    return df.drop(columns=["_completeness"])


def normalize_data(data: Dict) -> Dict:
    """
    Applies Colombian formatting rules to invoice data.
    """
    if not data:
        return {}

    # Ensure fecha is set from fecha_emision when missing (e.g. from XML)
    if not data.get("fecha") and data.get("fecha_emision"):
        data["fecha"] = data["fecha_emision"]

    # numeric fields
    for field in ["base", "impuestos", "total"]:
        if field in data:
            data[field] = clean_colombian_number(data[field])

    # Heuristic: correct suspiciously low total (decimal error)
    if "total" in data and data["total"] is not None:
        data["total"] = correct_suspicious_low_total(data["total"])

    # date fields
    for field in ["fecha", "fecha_vencimiento", "fecha_emision"]:
        if field in data:
            data[field] = format_date_colombian(data[field])

    # nit field
    if 'nit' in data and data['nit']:
        raw_nit = clean_nit(data['nit'])
        data['nit'] = raw_nit # Store the clean full string
        
        # Split into Number and DV
        # Possible formats: "123456-1", "123456"
        if '-' in raw_nit:
            parts = raw_nit.split('-')
            # Take the first part as number, last as DV
            # Handle cases like "123-456-1"? no, we removed dots.
            nit_num = parts[0]
            nit_dv_extracted = parts[-1]
        else:
            # If no hyphen, we assume it MIGHT be "1234561" where last is DV?
            # Or "123456" without DV?
            # The prompt says: "recibe nit sin puntos ni rayas para formato XXX.XXX.XXX - 9"
            # in valid_nit python function it treats last char as DV.
            # But here `clean_nit` preserves hyphens if they existed.
            # If the user OCR'd "900123456 1", clean_nit made it "9001234561".
            # This is ambiguous.
            # Generally, Vision API might return formatted "900.123.456-1".
            # If vision returned "900123456", we don't know if DV is included.
            # Lets assumption: 
            # If hyphen exists -> explicitly split.
            # If no hyphen -> assume NO DV extracted, or try to infer?
            # Safer: If no hyphen, leave 'nit_dv_extraido' as None.
            nit_num = raw_nit
            nit_dv_extracted = None
            
        data['nit'] = nit_num  # Update 'nit' to be just the number
        data['nit_dv_extraido'] = nit_dv_extracted
        
        # Calculate DV based on the number part
        if nit_num and nit_num.isdigit():
             data['nit_dv_calculado'] = calculate_nit_verification_digit(nit_num)
        else:
             data['nit_dv_calculado'] = "?"

    return data
